import os
import re
import PyPDF2
import pdfplumber
import pandas as pd
import logging
from typing import Dict, List, Any, Optional, Tuple, Union
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)

def find_benefit(text: str, keywords: List[str], context_chars: int = 200) -> str:
    """
    Find a benefit value in the text based on keywords.
    Returns the first match, surrounded by context.
    """
    text = text.lower()
    for keyword in keywords:
        pattern = r'(?i)(?:.*?)\b' + re.escape(keyword.lower()) + r'\b(.*?)(?:\$[\d,]+(?:\.\d+)?|(?:\d+)%|not covered|covered 100%)'
        match = re.search(pattern, text, re.DOTALL)
        if match:
            start_pos = max(0, match.start() - context_chars)
            end_pos = min(len(text), match.end() + context_chars)
            context = text[start_pos:end_pos]
            
            # Find the dollar amount or percentage in the context
            amount_match = re.search(r'\$[\d,]+(?:\.\d+)?|\d+%|covered 100%|not covered', context, re.IGNORECASE)
            if amount_match:
                return amount_match.group(0)
    
    return "Not found"

def find_percentage(text: str, keywords: List[str]) -> str:
    """
    Find a percentage value in the text based on keywords.
    """
    text = text.lower()
    for keyword in keywords:
        pattern = r'(?i)(?:.*?)\b' + re.escape(keyword.lower()) + r'\b.*?(\d+(?:\.\d+)?%)'
        match = re.search(pattern, text, re.DOTALL)
        if match:
            return match.group(1)
    
    return "Not found"

def create_benefit_excel(data: Dict[str, Any], output_path: str) -> bool:
    """Create an Excel file with the extracted benefits."""
    try:
        # Check if we're updating an existing template or creating a new file
        try:
            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["HEALTH"]
        except (FileNotFoundError, KeyError):
            # Create a new file with the required structure
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "HEALTH"
            
            # Set up basic structure and headers for the template
            sheet['A1'] = "ITEM"
            sheet['B1'] = "FEATURE"
            sheet['C1'] = "DESCRIPTION"
            sheet['D1'] = "Plan 1"
            sheet['E1'] = "In Network"
            sheet['F1'] = "Out of Network"
            sheet['G1'] = "Plan 2"
            sheet['H1'] = "In Network"
            sheet['I1'] = "Out of Network"
            
            # Set up row labels
            sheet['A4'] = "Carrier Name"
            sheet['A5'] = "Plan Name"
            sheet['A6'] = "Page Name"
            sheet['A7'] = "Plan Explanation"
            sheet['A9'] = "Single Deductible"
            sheet['A10'] = "Family Deductible"
            sheet['A12'] = "Coinsurance"
            sheet['A14'] = "Single Out of Pocket Max"
            sheet['A15'] = "Family Out of Pocket Max"
            sheet['A17'] = "Primary Care Office Visit"
            sheet['A18'] = "Specialist Office Visit"
            sheet['A19'] = "Urgent Care"
            sheet['A20'] = "Emergency Room"
            sheet['A22'] = "Preventive Services"
            sheet['A24'] = "Outpatient Surgery"
            sheet['A25'] = "Inpatient Hospitalization"
            sheet['A26'] = "CT Scan, PT Scan, MRI"
            sheet['A27'] = "Hospital Newborn Delivery"
            sheet['A29'] = "Prescription Deductible"
            sheet['A30'] = "Generic (Tier 1)"
            sheet['A31'] = "Brand Name (Tier 2)"
            sheet['A32'] = "Non-Preferred (Tier 3)"
            sheet['A33'] = "Specialty (Tier 4)"
            sheet['A34'] = "Specialty (Tier 5)"
            sheet['A35'] = "Mail Order (90 day supply)"
            sheet['A37'] = "Plan Year"
            sheet['A38'] = "Deductible Period"
            sheet['A39'] = "Deductible Explanation"
            sheet['A40'] = "Network Type"
            sheet['A41'] = "Network Name"
            sheet['A42'] = "Member Website"
            sheet['A43'] = "Customer Service Phone"
        
        # Find the next available plan column (D, G, J, etc.)
        next_plan_col = None
        for col_idx in range(4, sheet.max_column + 1, 3):  # Start from column D (index 4)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if not sheet[f'{col_letter}4'].value:  # Check carrier name cell
                next_plan_col = col_idx
                break
        
        if next_plan_col is None:
            # All slots are full, add a new plan column
            next_plan_col = sheet.max_column + 1
            
            # Add header for the new plan
            plan_header_cell = sheet.cell(row=1, column=next_plan_col)
            plan_header_cell.value = f"Plan {(next_plan_col - 1) // 3}"
            
            # Add In Network and Out of Network headers
            in_network_cell = sheet.cell(row=1, column=next_plan_col + 1)
            in_network_cell.value = "In Network"
            
            out_network_cell = sheet.cell(row=1, column=next_plan_col + 2)
            out_network_cell.value = "Out of Network"
        
        # Convert column index to letter
        col_letter = openpyxl.utils.get_column_letter(next_plan_col)
        in_col_letter = openpyxl.utils.get_column_letter(next_plan_col + 1)
        out_col_letter = openpyxl.utils.get_column_letter(next_plan_col + 2)
        
        # Get the results from the data
        result = data
        if "results" in data and len(data["results"]) > 0:
            result = data["results"][0]
        
        # Fill in plan information
        sheet[f'{col_letter}4'] = result.get("carrier_name", "Unknown")
        sheet[f'{col_letter}5'] = result.get("plan_name", "Unknown")
        sheet[f'{col_letter}6'] = "Health Insurance"
        sheet[f'{col_letter}7'] = "Health insurance provides financial protection against medical costs. It helps employees access necessary healthcare while minimizing out-of-pocket expenses."
        
        # Determine if plan is HSA
        is_hsa_plan = "HSA" in result.get("plan_name", "").upper()
        
        # Extract deductible values (numbers only for the spreadsheet)
        in_deductible = result.get("deductible", {}).get("individual_in_network", "")
        out_deductible = result.get("deductible", {}).get("individual_out_network", "")
        
        # Extract just the number from the deductible strings
        in_deductible_num = re.search(r'\$?([\d,]+)', in_deductible)
        out_deductible_num = re.search(r'\$?([\d,]+)', out_deductible)
        
        # Fill deductibles (numbers only)
        sheet[f'{in_col_letter}9'] = in_deductible_num.group(1).replace(',', '') if in_deductible_num else ""
        sheet[f'{out_col_letter}9'] = out_deductible_num.group(1).replace(',', '') if out_deductible_num else ""
        
        # Family deductibles
        in_family_deductible = result.get("deductible", {}).get("family_in_network", "")
        out_family_deductible = result.get("deductible", {}).get("family_out_network", "")
        
        # Extract just the number from the family deductible strings
        in_family_deductible_num = re.search(r'\$?([\d,]+)', in_family_deductible)
        out_family_deductible_num = re.search(r'\$?([\d,]+)', out_family_deductible)
        
        # Fill family deductibles (numbers only)
        sheet[f'{in_col_letter}10'] = in_family_deductible_num.group(1).replace(',', '') if in_family_deductible_num else ""
        sheet[f'{out_col_letter}10'] = out_family_deductible_num.group(1).replace(',', '') if out_family_deductible_num else ""
        
        # Coinsurance (numbers only)
        in_coinsurance = result.get("coinsurance", {}).get("in_network", "")
        out_coinsurance = result.get("coinsurance", {}).get("out_network", "")
        
        # Extract just the number from coinsurance strings
        in_coinsurance_num = re.search(r'(\d+)%', in_coinsurance)
        out_coinsurance_num = re.search(r'(\d+)%', out_coinsurance)
        
        # Fill coinsurance (numbers only)
        sheet[f'{in_col_letter}12'] = in_coinsurance_num.group(1) if in_coinsurance_num else ""
        sheet[f'{out_col_letter}12'] = out_coinsurance_num.group(1) if out_coinsurance_num else ""
        
        # Out of pocket max (numbers only)
        in_oop = result.get("out_of_pocket", {}).get("individual_in_network", "")
        out_oop = result.get("out_of_pocket", {}).get("individual_out_network", "")
        
        # Extract just the number from OOP strings
        in_oop_num = re.search(r'\$?([\d,]+)', in_oop)
        out_oop_num = re.search(r'\$?([\d,]+)', out_oop)
        
        # Fill OOP (numbers only)
        sheet[f'{in_col_letter}14'] = in_oop_num.group(1).replace(',', '') if in_oop_num else ""
        sheet[f'{out_col_letter}14'] = out_oop_num.group(1).replace(',', '') if out_oop_num else ""
        
        # Family OOP (numbers only)
        in_family_oop = result.get("out_of_pocket", {}).get("family_in_network", "")
        out_family_oop = result.get("out_of_pocket", {}).get("family_out_network", "")
        
        # Extract just the number from family OOP strings
        in_family_oop_num = re.search(r'\$?([\d,]+)', in_family_oop)
        out_family_oop_num = re.search(r'\$?([\d,]+)', out_family_oop)
        
        # Fill family OOP (numbers only)
        sheet[f'{in_col_letter}15'] = in_family_oop_num.group(1).replace(',', '') if in_family_oop_num else ""
        sheet[f'{out_col_letter}15'] = out_family_oop_num.group(1).replace(',', '') if out_family_oop_num else ""
        
        # Doctor visits - PCP
        primary_care = result.get("office_visits", {}).get("primary_care", "Not found")
        if primary_care != "Not found":
            # Format according to rules
            if re.match(r'^\$\d+$', primary_care):
                # For HSA plans, add "after deductible" to copay amounts
                if is_hsa_plan:
                    primary_care = f"{primary_care} after deductible"
            elif re.match(r'^\d+%$', primary_care):
                # Add "after deductible" to percentage amounts
                primary_care = f"{primary_care} after deductible"
        
        # Out of network PCP - usually from tables or estimated from coinsurance
        out_network_pcp = "50% after deductible"  # Default if not found
        
        sheet[f'{in_col_letter}17'] = primary_care
        sheet[f'{out_col_letter}17'] = out_network_pcp
        
        # Specialist visits
        specialist = result.get("office_visits", {}).get("specialist", "Not found")
        if specialist != "Not found":
            # Format according to rules
            if re.match(r'^\$\d+$', specialist):
                # For HSA plans, add "after deductible" to copay amounts
                if is_hsa_plan:
                    specialist = f"{specialist} after deductible"
            elif re.match(r'^\d+%$', specialist):
                # Add "after deductible" to percentage amounts
                specialist = f"{specialist} after deductible"
        
        # Out of network specialist
        out_network_specialist = "50% after deductible"  # Default if not found
        
        sheet[f'{in_col_letter}18'] = specialist
        sheet[f'{out_col_letter}18'] = out_network_specialist
        
        # Urgent care
        urgent_care = result.get("office_visits", {}).get("urgent_care", "Not found")
        if urgent_care != "Not found":
            # Format according to rules
            if re.match(r'^\$\d+$', urgent_care):
                # For HSA plans, add "after deductible" to copay amounts
                if is_hsa_plan:
                    urgent_care = f"{urgent_care} after deductible"
            elif re.match(r'^\d+%$', urgent_care):
                # Add "after deductible" to percentage amounts
                urgent_care = f"{urgent_care} after deductible"
        
        # Out of network urgent care
        out_network_urgent = "50% after deductible"  # Default if not found
        
        sheet[f'{in_col_letter}19'] = urgent_care
        sheet[f'{out_col_letter}19'] = out_network_urgent
        
        # Emergency room (same for in and out of network)
        emergency_room = result.get("emergency_room", "Not found")
        if emergency_room != "Not found":
            # Format according to rules
            if re.match(r'^\$\d+$', emergency_room):
                # For HSA plans, add "after deductible" to copay amounts
                if is_hsa_plan:
                    emergency_room = f"{emergency_room} after deductible"
            elif re.match(r'^\d+%$', emergency_room):
                # Add "after deductible" to percentage amounts
                emergency_room = f"{emergency_room} after deductible"
        
        sheet[f'{in_col_letter}20'] = emergency_room
        sheet[f'{out_col_letter}20'] = emergency_room  # Same as in-network per requirements
        
        # Preventive services (0% for in-network)
        sheet[f'{in_col_letter}22'] = "0%"
        sheet[f'{out_col_letter}22'] = "50% after deductible"  # Default if not found
        
        # Outpatient surgery
        outpatient_surgery = result.get("outpatient_surgery", "Not found")
        if outpatient_surgery != "Not found":
            # Check for freestanding vs hospital format
            if "Freestanding:" in outpatient_surgery:
                # Already formatted correctly
                pass
            else:
                # Format according to rules
                if re.match(r'^\$\d+$', outpatient_surgery):
                    # For HSA plans, add "after deductible" to copay amounts
                    if is_hsa_plan:
                        outpatient_surgery = f"{outpatient_surgery} after deductible"
                elif re.match(r'^\d+%$', outpatient_surgery):
                    # Add "after deductible" to percentage amounts
                    outpatient_surgery = f"{outpatient_surgery} after deductible"
        
        # Out of network outpatient surgery
        out_network_surgery = "50% after deductible"  # Default if not found
        
        sheet[f'{in_col_letter}24'] = outpatient_surgery
        sheet[f'{out_col_letter}24'] = out_network_surgery
        
        # Inpatient hospitalization
        hospitalization = result.get("hospitalization", "Not found")
        if hospitalization != "Not found":
            # Check for per occurrence format
            if "then" in hospitalization:
                # Already formatted correctly
                pass
            else:
                # Format according to rules
                if re.match(r'^\$\d+$', hospitalization):
                    # For HSA plans, add "after deductible" to copay amounts
                    if is_hsa_plan:
                        hospitalization = f"{hospitalization} after deductible"
                elif re.match(r'^\d+%$', hospitalization):
                    # Add "after deductible" to percentage amounts
                    hospitalization = f"{hospitalization} after deductible"
        
        # Out of network hospitalization
        out_network_hosp = "50% after deductible"  # Default if not found
        
        sheet[f'{in_col_letter}25'] = hospitalization
        sheet[f'{out_col_letter}25'] = out_network_hosp
        
        # CT/MRI/PT Scans
        imaging = result.get("imaging", "Not found")
        if imaging != "Not found":
            # Check for freestanding vs hospital format
            if "Freestanding:" in imaging:
                # Already formatted correctly
                pass
            else:
                # Format according to rules
                if re.match(r'^\$\d+$', imaging):
                    # For HSA plans, add "after deductible" to copay amounts
                    if is_hsa_plan:
                        imaging = f"{imaging} after deductible"
                elif re.match(r'^\d+%$', imaging):
                    # Add "after deductible" to percentage amounts
                    imaging = f"{imaging} after deductible"
        
        # Out of network imaging
        out_network_imaging = "50% after deductible"  # Default if not found
        
        sheet[f'{in_col_letter}26'] = imaging
        sheet[f'{out_col_letter}26'] = out_network_imaging
        
        # Hospital newborn delivery (same as hospitalization per requirements)
        sheet[f'{in_col_letter}27'] = hospitalization
        sheet[f'{out_col_letter}27'] = out_network_hosp
        
        # Prescription benefits
        if "prescription" in result:
            rx_benefits = result.get("prescription", {})
            
            # Prescription deductible
            rx_deductible = rx_benefits.get("deductible", "")
            if rx_deductible:
                sheet[f'{in_col_letter}29'] = rx_deductible
                sheet[f'{out_col_letter}29'] = rx_deductible
            
            # Generic (Tier 1)
            generic = rx_benefits.get("tier_1", "")
            if generic:
                if is_hsa_plan and re.match(r'^\$\d+$', generic):
                    generic = f"{generic} after deductible"
                sheet[f'{in_col_letter}30'] = generic
                sheet[f'{out_col_letter}30'] = "Not covered"
            
            # Brand (Tier 2)
            brand = rx_benefits.get("tier_2", "")
            if brand:
                if is_hsa_plan and re.match(r'^\$\d+$', brand):
                    brand = f"{brand} after deductible"
                sheet[f'{in_col_letter}31'] = brand
                sheet[f'{out_col_letter}31'] = "Not covered"
            
            # Non-preferred (Tier 3)
            non_preferred = rx_benefits.get("tier_3", "")
            if non_preferred:
                if is_hsa_plan and re.match(r'^\$\d+$', non_preferred):
                    non_preferred = f"{non_preferred} after deductible"
                sheet[f'{in_col_letter}32'] = non_preferred
                sheet[f'{out_col_letter}32'] = "Not covered"
            
            # Specialty (Tier 4)
            specialty = rx_benefits.get("tier_4", "")
            if specialty:
                if is_hsa_plan and re.match(r'^\$\d+$', specialty):
                    specialty = f"{specialty} after deductible"
                sheet[f'{in_col_letter}33'] = specialty
                sheet[f'{out_col_letter}33'] = "Not covered"
            
            # Specialty (Tier 5)
            specialty5 = rx_benefits.get("tier_5", "")
            if specialty5:
                if is_hsa_plan and re.match(r'^\$\d+$', specialty5):
                    specialty5 = f"{specialty5} after deductible"
                sheet[f'{in_col_letter}34'] = specialty5
                sheet[f'{out_col_letter}34'] = "Not covered"
            
            # Mail order
            mail_order = rx_benefits.get("mail_order", "")
            if mail_order:
                if is_hsa_plan and not "after deductible" in mail_order:
                    # Add to each cost in the tier list
                    mail_order_parts = mail_order.split("/")
                    formatted_parts = []
                    for part in mail_order_parts:
                        part = part.strip()
                        if re.match(r'^\$\d+$', part):
                            formatted_parts.append(f"{part} after deductible")
                        else:
                            formatted_parts.append(part)
                    mail_order = " / ".join(formatted_parts)
                sheet[f'{in_col_letter}35'] = mail_order
                sheet[f'{out_col_letter}35'] = "Not covered"
        
        # Additional plan information
        from datetime import datetime
        current_year = datetime.now().year
        
        sheet[f'{in_col_letter}37'] = str(current_year)
        sheet[f'{in_col_letter}38'] = "Calendar Year: January 1st – December 31st"
        
        # Deductible explanation
        deductible_type = "Embedded"  # Default
        if "Aggregate" in self.text or "aggregate" in self.text:
            deductible_type = "Aggregate"
        sheet[f'{in_col_letter}39'] = deductible_type
        
        # Network type
        network_type = "PPO"  # Default
        for network in ["PPO", "HMO", "EPO", "POS"]:
            if network in result.get("plan_name", "") or network in self.text:
                network_type = network
                break
        sheet[f'{in_col_letter}40'] = network_type
        
        # Network name (often next to network type)
        network_name_match = re.search(r'(?i)(?:network|network name|provider network)[\s:]+([A-Za-z0-9\s\-]+)', self.text)
        network_name = network_name_match.group(1).strip() if network_name_match else ""
        sheet[f'{in_col_letter}41'] = network_name
        
        # Member website
        website_match = re.search(r'(?i)(?:website|member website|web)[\s:]+([A-Za-z0-9\s\-\.\/:]+\.[a-z]{2,})', self.text)
        website = website_match.group(1).strip() if website_match else ""
        if not website:
            # Try to find URLs
            url_match = re.search(r'(?i)((?:https?:\/\/)?(?:www\.)?[a-zA-Z0-9-]+(?:\.[a-zA-Z]{2,})+(?:\/[a-zA-Z0-9-_.]+)*)', self.text)
            website = url_match.group(1) if url_match else ""
        sheet[f'{in_col_letter}42'] = website
        
        # Customer service phone
        phone_match = re.search(r'(?i)(?:customer service|phone|call|contact)[\s:]+(?:\d{1,2}-?)?(?:\d{3})[\s\.-]+(?:\d{3})[\s\.-]+(?:\d{4})', self.text)
        phone = phone_match.group(0) if phone_match else ""
        if not phone:
            # Try a simpler pattern for phone numbers
            phone_match = re.search(r'(?i)(?:\d{1,2}-?)?(?:\d{3})[\s\.-]+(?:\d{3})[\s\.-]+(?:\d{4})', self.text)
            phone = phone_match.group(0) if phone_match else ""
        
        # Format the phone number
        if phone:
            # Extract just the digits
            digits = re.findall(r'\d', phone)
            if len(digits) >= 10:
                # Format as XXX-XXX-XXXX
                start_idx = 0 if len(digits) == 10 else len(digits) - 10
                phone = f"{digits[start_idx]}{digits[start_idx+1]}{digits[start_idx+2]}-{digits[start_idx+3]}{digits[start_idx+4]}{digits[start_idx+5]}-{digits[start_idx+6]}{digits[start_idx+7]}{digits[start_idx+8]}{digits[start_idx+9]}"
        
        sheet[f'{in_col_letter}43'] = phone
        
        # Save workbook
        workbook.save(output_path)
        return True
    
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        return False

class BenefitExtractor:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.text = ""
        self.tables = []
        self.benefits = {}
    
    def extract_benefits(self) -> Dict[str, Any]:
        """Extract all benefits information from the PDF."""
        self._extract_text()
        self._extract_tables()
        self._extract_benefits()
        return self.get_formatted_benefits()
    
    def _extract_text(self) -> str:
        """Extract text from the PDF."""
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n\n"
                self.text = text
                return text
        except Exception as e:
            logger.error(f"Error extracting text: {str(e)}")
            # Fallback to PyPDF2
            try:
                with open(self.pdf_path, 'rb') as file:
                    reader = PyPDF2.PdfReader(file)
                    text = ""
                    for page in reader.pages:
                        text += page.extract_text() + "\n\n"
                    self.text = text
                    return text
            except Exception as e2:
                logger.error(f"PyPDF2 extraction failed: {str(e2)}")
                raise Exception(f"Failed to extract text from PDF: {str(e)}, {str(e2)}")
    
    def _extract_tables(self) -> List[Dict[str, Any]]:
        """Extract tables from the PDF."""
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                tables = []
                for i, page in enumerate(pdf.pages):
                    page_tables = page.extract_tables()
                    for j, table_data in enumerate(page_tables):
                        if table_data and len(table_data) > 0:
                            tables.append({
                                'page': i + 1,
                                'table_number': j + 1,
                                'data': table_data
                            })
                self.tables = tables
                return tables
        except Exception as e:
            logger.error(f"Error extracting tables: {str(e)}")
            self.tables = []
            return []
    
    def _extract_benefits(self) -> Dict[str, Any]:
        """Extract specific benefit information from text and tables."""
        # Extract carrier and plan information
        carrier_name = self._extract_carrier_name()
        plan_name = self._extract_plan_name()
        
        # Extract deductible information
        deductible = self._extract_deductible_info()
        
        # Extract out-of-pocket information
        out_of_pocket = self._extract_out_of_pocket_info()
        
        # Extract coinsurance
        coinsurance = self._extract_coinsurance_info()
        
        # Extract office visit information
        office_visits = self._extract_office_visit_info()
        
        # Extract emergency room information
        emergency_room = self._extract_emergency_room_info()
        
        # Extract hospitalization information
        hospitalization = self._extract_hospitalization_info()
        
        # Compile all benefits
        self.benefits = {
            "carrier_name": carrier_name,
            "plan_name": plan_name,
            "deductible": deductible,
            "out_of_pocket": out_of_pocket,
            "coinsurance": coinsurance,
            "office_visits": office_visits,
            "emergency_room": emergency_room,
            "hospitalization": hospitalization
        }
        
        return self.benefits
    
    def _extract_carrier_name(self) -> str:
        """Extract the carrier name from the PDF."""
        carrier_patterns = [
            r'(?i)(United\s*Healthcare|Aetna|Cigna|Blue\s*Cross|Blue\s*Shield|BCBS|Anthem|Humana|Kaiser|Optum)'
        ]
        
        for pattern in carrier_patterns:
            match = re.search(pattern, self.text)
            if match:
                return match.group(1)
        
        # Check the first tables for carrier info
        if self.tables:
            for table in self.tables[:2]:  # Only check first two tables
                for row in table['data']:
                    for cell in row:
                        if cell and isinstance(cell, str):
                            for pattern in carrier_patterns:
                                match = re.search(pattern, cell)
                                if match:
                                    return match.group(1)
        
        return "Unknown"
    
    def _extract_plan_name(self) -> str:
        """Extract the plan name from the PDF."""
        plan_patterns = [
            r'(?i)Plan\s*(?:Name|Type):\s*([A-Za-z0-9\s\-]+)',
            r'(?i)(Choice\s*(?:Plus|Select)|PPO|HMO|EPO|POS|HDHP|HSA)'
        ]
        
        for pattern in plan_patterns:
            match = re.search(pattern, self.text)
            if match:
                return match.group(1).strip()
        
        return "Unknown"
    
    def _extract_deductible_info(self) -> Dict[str, str]:
        """Extract deductible information."""
        # Regular expressions for deductible amounts
        individual_in_pattern = r'(?i)Individual[^\n]*(?:In[\s\-]*Network|Network)[^\n]*?(\$[\d,]+)'
        family_in_pattern = r'(?i)Family[^\n]*(?:In[\s\-]*Network|Network)[^\n]*?(\$[\d,]+)'
        individual_out_pattern = r'(?i)Individual[^\n]*(?:Out[\s\-]*of[\s\-]*Network|Non[\s\-]*Network)[^\n]*?(\$[\d,]+)'
        family_out_pattern = r'(?i)Family[^\n]*(?:Out[\s\-]*of[\s\-]*Network|Non[\s\-]*Network)[^\n]*?(\$[\d,]+)'
        
        # Simple pattern fallbacks
        fallback_individual = r'(?i)Individual\s*(?:Deductible|Annual\s*Deductible)[^\n]*?(\$[\d,]+)'
        fallback_family = r'(?i)Family\s*(?:Deductible|Annual\s*Deductible)[^\n]*?(\$[\d,]+)'
        
        # Try to match patterns
        individual_in = "Not found"
        match = re.search(individual_in_pattern, self.text)
        if match:
            individual_in = match.group(1)
        else:
            match = re.search(fallback_individual, self.text)
            if match:
                individual_in = match.group(1)
        
        family_in = "Not found"
        match = re.search(family_in_pattern, self.text)
        if match:
            family_in = match.group(1)
        else:
            match = re.search(fallback_family, self.text)
            if match:
                family_in = match.group(1)
        
        individual_out = "Not found"
        match = re.search(individual_out_pattern, self.text)
        if match:
            individual_out = match.group(1)
        
        family_out = "Not found"
        match = re.search(family_out_pattern, self.text)
        if match:
            family_out = match.group(1)
        
        # Check tables if text search failed
        if individual_in == "Not found" or family_in == "Not found":
            for table in self.tables:
                for row in table['data']:
                    if not all(row):
                        continue
                    row_text = ' '.join([str(cell) for cell in row if cell])
                    if 'deductible' in row_text.lower() and 'individual' in row_text.lower():
                        for cell in row:
                            if cell and isinstance(cell, str) and '$' in cell:
                                individual_in = cell
                                break
                    if 'deductible' in row_text.lower() and 'family' in row_text.lower():
                        for cell in row:
                            if cell and isinstance(cell, str) and '$' in cell:
                                family_in = cell
                                break
        
        return {
            "individual_in_network": individual_in,
            "family_in_network": family_in,
            "individual_out_network": individual_out,
            "family_out_network": family_out
        }
    
    def _extract_out_of_pocket_info(self) -> Dict[str, str]:
        """Extract out-of-pocket information."""
        # Regular expressions for out-of-pocket amounts
        individual_in_pattern = r'(?i)(?:Individual|Out[\s\-]*of[\s\-]*Pocket\s*(?:Limit|Maximum))[^\n]*(?:In[\s\-]*Network|Network)[^\n]*?(\$[\d,]+)'
        family_in_pattern = r'(?i)(?:Family|Out[\s\-]*of[\s\-]*Pocket\s*(?:Limit|Maximum))[^\n]*(?:In[\s\-]*Network|Network)[^\n]*?(\$[\d,]+)'
        individual_out_pattern = r'(?i)(?:Individual|Out[\s\-]*of[\s\-]*Pocket\s*(?:Limit|Maximum))[^\n]*(?:Out[\s\-]*of[\s\-]*Network|Non[\s\-]*Network)[^\n]*?(\$[\d,]+)'
        family_out_pattern = r'(?i)(?:Family|Out[\s\-]*of[\s\-]*Pocket\s*(?:Limit|Maximum))[^\n]*(?:Out[\s\-]*of[\s\-]*Network|Non[\s\-]*Network)[^\n]*?(\$[\d,]+)'
        
        # Simple pattern fallbacks
        fallback_individual = r'(?i)Individual\s*(?:Out[\s\-]*of[\s\-]*Pocket|OOP)[^\n]*?(\$[\d,]+)'
        fallback_family = r'(?i)Family\s*(?:Out[\s\-]*of[\s\-]*Pocket|OOP)[^\n]*?(\$[\d,]+)'
        
        # Try to match patterns
        individual_in = "Not found"
        match = re.search(individual_in_pattern, self.text)
        if match:
            individual_in = match.group(1)
        else:
            match = re.search(fallback_individual, self.text)
            if match:
                individual_in = match.group(1)
        
        family_in = "Not found"
        match = re.search(family_in_pattern, self.text)
        if match:
            family_in = match.group(1)
        else:
            match = re.search(fallback_family, self.text)
            if match:
                family_in = match.group(1)
        
        individual_out = "Not found"
        match = re.search(individual_out_pattern, self.text)
        if match:
            individual_out = match.group(1)
        
        family_out = "Not found"
        match = re.search(family_out_pattern, self.text)
        if match:
            family_out = match.group(1)
        
        # Check tables
        if individual_in == "Not found" or family_in == "Not found":
            for table in self.tables:
                for row in table['data']:
                    if not all(row):
                        continue
                    row_text = ' '.join([str(cell) for cell in row if cell])
                    if any(phrase in row_text.lower() for phrase in ['out-of-pocket', 'out of pocket', 'oop']) and 'individual' in row_text.lower():
                        for cell in row:
                            if cell and isinstance(cell, str) and '$' in cell:
                                individual_in = cell
                                break
                    if any(phrase in row_text.lower() for phrase in ['out-of-pocket', 'out of pocket', 'oop']) and 'family' in row_text.lower():
                        for cell in row:
                            if cell and isinstance(cell, str) and '$' in cell:
                                family_in = cell
                                break
        
        return {
            "individual_in_network": individual_in,
            "family_in_network": family_in,
            "individual_out_network": individual_out,
            "family_out_network": family_out
        }
    
    def _extract_coinsurance_info(self) -> Dict[str, str]:
        """Extract coinsurance information."""
        in_network_pattern = r'(?i)(?:Co(?:\-|\s)?insurance)[^\n]*(?:In[\s\-]*Network|Network)[^\n]*?(\d+(?:\.\d+)?%)'
        out_network_pattern = r'(?i)(?:Co(?:\-|\s)?insurance)[^\n]*(?:Out[\s\-]*of[\s\-]*Network|Non[\s\-]*Network)[^\n]*?(\d+(?:\.\d+)?%)'
        
        # Fallback pattern
        fallback_pattern = r'(?i)Co(?:\-|\s)?insurance(?:[^\n]*?)(\d+(?:\.\d+)?%)'
        
        # Try to match patterns
        in_network = "Not found"
        match = re.search(in_network_pattern, self.text)
        if match:
            in_network = match.group(1)
        else:
            match = re.search(fallback_pattern, self.text)
            if match:
                in_network = match.group(1)
        
        out_network = "Not found"
        match = re.search(out_network_pattern, self.text)
        if match:
            out_network = match.group(1)
        
        # Check tables
        if in_network == "Not found":
            for table in self.tables:
                for row in table['data']:
                    if not all(row):
                        continue
                    row_text = ' '.join([str(cell) for cell in row if cell])
                    if 'coinsurance' in row_text.lower():
                        for cell in row:
                            if cell and isinstance(cell, str) and '%' in cell:
                                in_network = cell
                                break
        
        return {
            "in_network": in_network,
            "out_network": out_network
        }
    
    def _extract_office_visit_info(self) -> Dict[str, str]:
        """Extract office visit information."""
        primary_care_pattern = r'(?i)(?:Primary\s*Care|PCP)[^\n]*(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)'
        specialist_pattern = r'(?i)(?:Specialist)[^\n]*(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)'
        urgent_care_pattern = r'(?i)(?:Urgent\s*Care)[^\n]*(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)'
        
        # Extract primary care
        primary_care = "Not found"
        match = re.search(primary_care_pattern, self.text)
        if match:
            cost_match = re.search(r'(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)', match.group(0))
            if cost_match:
                primary_care = cost_match.group(0)
        
        # Extract specialist
        specialist = "Not found"
        match = re.search(specialist_pattern, self.text)
        if match:
            cost_match = re.search(r'(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)', match.group(0))
            if cost_match:
                specialist = cost_match.group(0)
        
        # Extract urgent care
        urgent_care = "Not found"
        match = re.search(urgent_care_pattern, self.text)
        if match:
            cost_match = re.search(r'(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)', match.group(0))
            if cost_match:
                urgent_care = cost_match.group(0)
        
        # Check tables
        if primary_care == "Not found" or specialist == "Not found" or urgent_care == "Not found":
            for table in self.tables:
                for row in table['data']:
                    if not all(row):
                        continue
                    row_text = ' '.join([str(cell) for cell in row if cell])
                    
                    if primary_care == "Not found" and any(term in row_text.lower() for term in ['primary care', 'pcp']):
                        for cell in row:
                            if cell and isinstance(cell, str) and re.search(r'(?:\$[\d,]+|\d+%)', cell):
                                primary_care = cell
                                break
                    
                    if specialist == "Not found" and 'specialist' in row_text.lower():
                        for cell in row:
                            if cell and isinstance(cell, str) and re.search(r'(?:\$[\d,]+|\d+%)', cell):
                                specialist = cell
                                break
                                
                    if urgent_care == "Not found" and 'urgent care' in row_text.lower():
                        for cell in row:
                            if cell and isinstance(cell, str) and re.search(r'(?:\$[\d,]+|\d+%)', cell):
                                urgent_care = cell
                                break
        
        return {
            "primary_care": primary_care,
            "specialist": specialist,
            "urgent_care": urgent_care
        }
    
    def _extract_emergency_room_info(self) -> str:
        """Extract emergency room information."""
        er_pattern = r'(?i)(?:Emergency\s*(?:Room|Department|Care|Services)|ER\s*Visit)[^\n]*(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)'
        
        # Extract emergency room
        er_cost = "Not found"
        match = re.search(er_pattern, self.text)
        if match:
            cost_match = re.search(r'(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)', match.group(0))
            if cost_match:
                er_cost = cost_match.group(0)
        
        # Check tables
        if er_cost == "Not found":
            for table in self.tables:
                for row in table['data']:
                    if not all(row):
                        continue
                    row_text = ' '.join([str(cell) for cell in row if cell])
                    if any(term in row_text.lower() for term in ['emergency', 'er visit']):
                        for cell in row:
                            if cell and isinstance(cell, str) and re.search(r'(?:\$[\d,]+|\d+%)', cell):
                                er_cost = cell
                                break
        
        return er_cost
    
    def _extract_hospitalization_info(self) -> str:
        """Extract hospitalization information."""
        hospitalization_pattern = r'(?i)(?:Hospital|Inpatient|Hospitalization)[^\n]*(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)'
        
        # Extract hospitalization
        hospitalization_cost = "Not found"
        match = re.search(hospitalization_pattern, self.text)
        if match:
            cost_match = re.search(r'(?:\$[\d,]+(?:\.\d+)?|\d+%|not covered|covered 100%)', match.group(0))
            if cost_match:
                hospitalization_cost = cost_match.group(0)
        
        # Check tables
        if hospitalization_cost == "Not found":
            for table in self.tables:
                for row in table['data']:
                    if not all(row):
                        continue
                    row_text = ' '.join([str(cell) for cell in row if cell])
                    if any(term in row_text.lower() for term in ['hospital', 'inpatient', 'hospitalization']):
                        for cell in row:
                            if cell and isinstance(cell, str) and re.search(r'(?:\$[\d,]+|\d+%)', cell):
                                hospitalization_cost = cell
                                break
        
        return hospitalization_cost
    
    def get_formatted_benefits(self) -> Dict[str, Any]:
        """Get the formatted benefits for display or export."""
        # If benefits haven't been extracted yet, do it now
        if not self.benefits:
            self._extract_benefits()
        
        # Apply formatting rules based on plan type and special formatting requirements
        formatted_benefits = self.benefits.copy()
        
        # Check if this is an HSA plan
        is_hsa_plan = "HSA" in formatted_benefits.get("plan_name", "").upper()
        
        # Format office visits based on HSA plan rules
        if "office_visits" in formatted_benefits:
            office_visits = formatted_benefits["office_visits"]
            
            # Primary care visit formatting
            if "primary_care" in office_visits:
                primary_care = office_visits["primary_care"]
                if re.match(r'^\$\d+$', primary_care):
                    # For HSA plans, add "after deductible" to copay amounts
                    if is_hsa_plan:
                        office_visits["primary_care"] = f"{primary_care} after deductible"
                elif re.match(r'^\d+%$', primary_care):
                    # Add "after deductible" to percentage amounts
                    office_visits["primary_care"] = f"{primary_care} after deductible"
            
            # Specialist visit formatting
            if "specialist" in office_visits:
                specialist = office_visits["specialist"]
                if re.match(r'^\$\d+$', specialist):
                    # For HSA plans, add "after deductible" to copay amounts
                    if is_hsa_plan:
                        office_visits["specialist"] = f"{specialist} after deductible"
                elif re.match(r'^\d+%$', specialist):
                    # Add "after deductible" to percentage amounts
                    office_visits["specialist"] = f"{specialist} after deductible"
            
            # Urgent care formatting
            if "urgent_care" in office_visits:
                urgent_care = office_visits["urgent_care"]
                if re.match(r'^\$\d+$', urgent_care):
                    # For HSA plans, add "after deductible" to copay amounts
                    if is_hsa_plan:
                        office_visits["urgent_care"] = f"{urgent_care} after deductible"
                elif re.match(r'^\d+%$', urgent_care):
                    # Add "after deductible" to percentage amounts
                    office_visits["urgent_care"] = f"{urgent_care} after deductible"
        
        # Format emergency room
        if "emergency_room" in formatted_benefits:
            emergency_room = formatted_benefits["emergency_room"]
            if re.match(r'^\$\d+$', emergency_room):
                # For HSA plans, add "after deductible" to copay amounts
                if is_hsa_plan:
                    formatted_benefits["emergency_room"] = f"{emergency_room} after deductible"
            elif re.match(r'^\d+%$', emergency_room):
                # Add "after deductible" to percentage amounts
                formatted_benefits["emergency_room"] = f"{emergency_room} after deductible"
        
        # Format hospitalization - check for per occurrence deductibles
        if "hospitalization" in formatted_benefits:
            hospitalization = formatted_benefits["hospitalization"]
            
            # Check for per occurrence deductible
            per_occurrence_match = re.search(r'(?i)(\$\d+)\s+per\s+(?:occurrence|admission)', hospitalization)
            percentage_match = re.search(r'(\d+)%', hospitalization)
            
            if per_occurrence_match and percentage_match:
                per_occurrence = per_occurrence_match.group(1)
                percentage = percentage_match.group(1)
                formatted_benefits["hospitalization"] = f"{per_occurrence}, then {percentage}% after deductible"
            elif re.match(r'^\d+%$', hospitalization):
                # Add "after deductible" to percentage amounts
                formatted_benefits["hospitalization"] = f"{hospitalization} after deductible"
        
        # Add multi-level formatting for facilities (e.g., outpatient/inpatient surgery)
        if "outpatient_surgery" in formatted_benefits:
            # Check for different facility types
            freestanding_match = re.search(r'(?i)(?:Freestanding|Ambulatory|Free\s+standing).*?(\d+%|\$\d+)', self.text)
            hospital_match = re.search(r'(?i)(?:Hospital|Outpatient\s+hospital).*?facility.*?(\d+%|\$\d+)', self.text)
            
            if freestanding_match and hospital_match:
                freestanding_value = freestanding_match.group(1)
                hospital_value = hospital_match.group(1)
                
                # Format each value properly
                if re.match(r'^\d+%$', freestanding_value):
                    freestanding_value += " after deductible"
                elif re.match(r'^\$\d+$', freestanding_value) and is_hsa_plan:
                    freestanding_value += " after deductible"
                    
                if re.match(r'^\d+%$', hospital_value):
                    hospital_value += " after deductible"
                elif re.match(r'^\$\d+$', hospital_value) and is_hsa_plan:
                    hospital_value += " after deductible"
                
                formatted_benefits["outpatient_surgery"] = f"Freestanding: {freestanding_value} / Hospital: {hospital_value}"
        
        # For imaging services (CT, MRI, etc.)
        if "imaging" in formatted_benefits:
            # Check for different facility types
            freestanding_match = re.search(r'(?i)(?:Freestanding|Ambulatory|Free\s+standing).*?(?:scan|imaging|diagnostic).*?(\d+%|\$\d+)', self.text)
            hospital_match = re.search(r'(?i)(?:Hospital|Outpatient\s+hospital).*?(?:scan|imaging|diagnostic).*?(\d+%|\$\d+)', self.text)
            
            if freestanding_match and hospital_match:
                freestanding_value = freestanding_match.group(1)
                hospital_value = hospital_match.group(1)
                
                # Format each value properly
                if re.match(r'^\d+%$', freestanding_value):
                    freestanding_value += " after deductible"
                elif re.match(r'^\$\d+$', freestanding_value) and is_hsa_plan:
                    freestanding_value += " after deductible"
                    
                if re.match(r'^\d+%$', hospital_value):
                    hospital_value += " after deductible"
                elif re.match(r'^\$\d+$', hospital_value) and is_hsa_plan:
                    hospital_value += " after deductible"
                
                formatted_benefits["imaging"] = f"Freestanding: {freestanding_value} / Hospital: {hospital_value}"
        
        # Format prescription drug benefits - check for specialty medications
        if "prescription" in formatted_benefits:
            rx_benefits = formatted_benefits["prescription"]
            
            # Format each tier
            for tier_key in rx_benefits:
                if "tier" in tier_key.lower():
                    tier_value = rx_benefits[tier_key]
                    
                    # Check for specialty notation
                    specialty_match = re.search(r'(?i)specialty.*?(\$\d+|\d+%)', self.text)
                    if specialty_match and "tier" in tier_key.lower() and ("specialty" in tier_key.lower() or int(tier_key[-1]) >= 3):
                        specialty_value = specialty_match.group(1)
                        
                        # Format specialty value
                        if re.match(r'^\d+%$', specialty_value):
                            specialty_value += " after deductible"
                        elif re.match(r'^\$\d+$', specialty_value) and is_hsa_plan:
                            specialty_value += " after deductible"
                            
                        # Add specialty notation
                        rx_benefits[tier_key] = f"{tier_value} (Specialty: {specialty_value})"
                    
                    # Format regular tier value
                    elif re.match(r'^\$\d+$', tier_value) and is_hsa_plan:
                        rx_benefits[tier_key] = f"{tier_value} after deductible"
                    elif re.match(r'^\d+%$', tier_value):
                        rx_benefits[tier_key] = f"{tier_value} after deductible"
        
        # Set preventive services to 0% for in-network
        if "preventive_services" not in formatted_benefits:
            formatted_benefits["preventive_services"] = {"in_network": "0%", "out_network": "Not found"}
        else:
            formatted_benefits["preventive_services"]["in_network"] = "0%"
        
        return formatted_benefits
