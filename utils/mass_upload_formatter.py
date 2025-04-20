import re
import logging
import openpyxl
from typing import Dict, List, Any, Optional, Tuple, Union
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)

def format_benefit_excel(data: Dict[str, Any], output_path: str) -> bool:
    """
    Create an Excel file with the extracted benefits, formatted specifically 
    for the mass upload template requirements.
    
    Special Formatting Rules:
    1. For non-HSA plans, simple copay amounts should just be the dollar amount (e.g., "$15")
    2. For HSA plans, add "after deductible" after dollar amounts (e.g., "$15 after deductible")
    3. For percentage values, always add "after deductible" (e.g., "20% after deductible")
    4. For services with different costs at different facilities, format as "Freestanding: X / Hospital: Y"
    5. For per occurrence deductibles, format as "$X, then Y% after deductible"
    6. For mail order prescriptions with multiple tiers, format as "Tier1 / Tier2 / Tier3" (e.g., "$10 / $20 / $40")
    7. For preventive services in-network, always use "0%"
    8. Emergency room out-of-network should match the in-network value
    9. Hospital newborn delivery should match inpatient hospitalization
    """
    try:
        # Check if we're updating an existing template or creating a new file
        try:
            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook["HEALTH"]
        except (FileNotFoundError, KeyError):
            # Create a new file with the required structure
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "HEALTH"
            
            # Set up basic structure and headers for the template
            sheet['A1'] = "ITEM"
            sheet['B1'] = "FEATURE"
            sheet['C1'] = "DESCRIPTION"
            sheet['D1'] = "Plan 1"
            sheet['E1'] = "In Network"
            sheet['F1'] = "Out of Network"
            sheet['G1'] = "Plan 2"
            sheet['H1'] = "In Network"
            sheet['I1'] = "Out of Network"
            
            # Set up row labels
            sheet['A4'] = "Carrier Name"
            sheet['A5'] = "Plan Name"
            sheet['A6'] = "Page Name"
            sheet['A7'] = "Plan Explanation"
            sheet['A9'] = "Single Deductible"
            sheet['A10'] = "Family Deductible"
            sheet['A12'] = "Coinsurance"
            sheet['A14'] = "Single Out of Pocket Max"
            sheet['A15'] = "Family Out of Pocket Max"
            sheet['A17'] = "Primary Care Office Visit"
            sheet['A18'] = "Specialist Office Visit"
            sheet['A19'] = "Urgent Care"
            sheet['A20'] = "Emergency Room"
            sheet['A22'] = "Preventive Services"
            sheet['A24'] = "Outpatient Surgery"
            sheet['A25'] = "Inpatient Hospitalization"
            sheet['A26'] = "CT Scan, PT Scan, MRI"
            sheet['A27'] = "Hospital Newborn Delivery"
            sheet['A29'] = "Prescription Deductible"
            sheet['A30'] = "Generic (Tier 1)"
            sheet['A31'] = "Brand Name (Tier 2)"
            sheet['A32'] = "Non-Preferred (Tier 3)"
            sheet['A33'] = "Specialty (Tier 4)"
            sheet['A34'] = "Specialty (Tier 5)"
            sheet['A35'] = "Mail Order (90 day supply)"
            sheet['A37'] = "Plan Year"
            sheet['A38'] = "Deductible Period"
            sheet['A39'] = "Deductible Explanation"
            sheet['A40'] = "Network Type"
            sheet['A41'] = "Network Name"
            sheet['A42'] = "Member Website"
            sheet['A43'] = "Customer Service Phone"
        
        # Find the next available plan column (D, G, J, etc.)
        next_plan_col = None
        for col_idx in range(4, sheet.max_column + 1, 3):  # Start from column D (index 4)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if not sheet[f'{col_letter}4'].value:  # Check carrier name cell
                next_plan_col = col_idx
                break
        
        if next_plan_col is None:
            # All slots are full, add a new plan column
            next_plan_col = sheet.max_column + 1
            
            # Add header for the new plan
            plan_header_cell = sheet.cell(row=1, column=next_plan_col)
            plan_header_cell.value = f"Plan {(next_plan_col - 1) // 3}"
            
            # Add In Network and Out of Network headers
            in_network_cell = sheet.cell(row=1, column=next_plan_col + 1)
            in_network_cell.value = "In Network"
            
            out_network_cell = sheet.cell(row=1, column=next_plan_col + 2)
            out_network_cell.value = "Out of Network"
        
        # Convert column index to letter
        col_letter = openpyxl.utils.get_column_letter(next_plan_col)
        in_col_letter = openpyxl.utils.get_column_letter(next_plan_col + 1)
        out_col_letter = openpyxl.utils.get_column_letter(next_plan_col + 2)
        
        # Get the results from the data - handle both direct data and nested results
        result = data
        if isinstance(data, dict) and "results" in data and isinstance(data["results"], list) and len(data["results"]) > 0:
            result = data["results"][0]
        
        # Fill in plan information
        sheet[f'{col_letter}4'] = result.get("carrier_name", "Unknown")
        sheet[f'{col_letter}5'] = result.get("plan_name", "Unknown")
        sheet[f'{col_letter}6'] = "Health Insurance"
        sheet[f'{col_letter}7'] = "Health insurance provides financial protection against medical costs. It helps employees access necessary healthcare while minimizing out-of-pocket expenses."
        
        # Determine if plan is HSA
        plan_name = result.get("plan_name", "").upper()
        is_hsa_plan = any(term in plan_name for term in ["HSA", "HEALTH SAVINGS", "HIGH DEDUCTIBLE"])
        
        # Helper function to format benefit values according to rules
        def format_benefit(value, is_percentage=False):
            """Format benefit value according to formatting rules"""
            if not value or value == "Not found":
                return ""
                
            # Already formatted with facility differences
            if "freestanding" in str(value).lower() or "hospital" in str(value).lower():
                return value
                
            # Already formatted with per occurrence deductible
            if "$" in str(value) and "then" in str(value).lower():
                if "after deductible" not in str(value).lower() and is_percentage:
                    return f"{value} after deductible"
                return value
                
            # Regular copay amount
            if re.match(r'^\$\d+(?:\.\d+)?$', str(value)):
                if is_hsa_plan and "after deductible" not in str(value).lower():
                    return f"{value} after deductible"
                return value
                
            # Percentage value
            if re.match(r'^\d+(?:\.\d+)?%$', str(value)):
                if "after deductible" not in str(value).lower():
                    return f"{value} after deductible"
                return value
                
            return value
            
        # Helper function to extract number only from monetary values
        def extract_number(value):
            """Extract just the numeric portion from monetary values"""
            if not value or value == "Not found":
                return ""
                
            match = re.search(r'\$?([\d,]+(?:\.\d+)?)', str(value))
            if match:
                return match.group(1).replace(',', '')
            return ""
            
        # Helper function to extract percentage only
        def extract_percentage(value):
            """Extract just the percentage number"""
            if not value or value == "Not found":
                return ""
                
            match = re.search(r'(\d+(?:\.\d+)?)%', str(value))
            if match:
                return match.group(1)
            return ""
        
        # ----- DEDUCTIBLES -----
        # In-network individual deductible
        in_deductible = result.get("deductible", {}).get("individual_in_network", "")
        sheet[f'{in_col_letter}9'] = extract_number(in_deductible)
        
        # Out-of-network individual deductible
        out_deductible = result.get("deductible", {}).get("individual_out_network", "")
        sheet[f'{out_col_letter}9'] = extract_number(out_deductible)
        
        # In-network family deductible
        in_family_deductible = result.get("deductible", {}).get("family_in_network", "")
        sheet[f'{in_col_letter}10'] = extract_number(in_family_deductible)
        
        # Out-of-network family deductible
        out_family_deductible = result.get("deductible", {}).get("family_out_network", "")
        sheet[f'{out_col_letter}10'] = extract_number(out_family_deductible)
        
        # ----- COINSURANCE -----
        in_coinsurance = result.get("coinsurance", {}).get("in_network", "")
        sheet[f'{in_col_letter}12'] = extract_percentage(in_coinsurance)
        
        out_coinsurance = result.get("coinsurance", {}).get("out_network", "")
        sheet[f'{out_col_letter}12'] = extract_percentage(out_coinsurance)
        
        # ----- OUT OF POCKET MAXIMUMS -----
        in_oop = result.get("out_of_pocket", {}).get("individual_in_network", "")
        sheet[f'{in_col_letter}14'] = extract_number(in_oop)
        
        out_oop = result.get("out_of_pocket", {}).get("individual_out_network", "")
        sheet[f'{out_col_letter}14'] = extract_number(out_oop)
        
        in_family_oop = result.get("out_of_pocket", {}).get("family_in_network", "")
        sheet[f'{in_col_letter}15'] = extract_number(in_family_oop)
        
        out_family_oop = result.get("out_of_pocket", {}).get("family_out_network", "")
        sheet[f'{out_col_letter}15'] = extract_number(out_family_oop)
        
        # ----- OFFICE VISITS -----
        # Primary Care
        primary_care = result.get("office_visits", {}).get("primary_care", "Not found")
        primary_care = format_benefit(primary_care)
        sheet[f'{in_col_letter}17'] = primary_care
        
        # Out-of-network PCP
        out_network_pcp = result.get("office_visits", {}).get("primary_care_out_network", "Not found")
        if out_network_pcp == "Not found" and out_coinsurance:
            out_network_pcp = f"{extract_percentage(out_coinsurance)}% after deductible"
        sheet[f'{out_col_letter}17'] = out_network_pcp
        
        # Specialist
        specialist = result.get("office_visits", {}).get("specialist", "Not found")
        specialist = format_benefit(specialist)
        sheet[f'{in_col_letter}18'] = specialist
        
        # Out-of-network Specialist
        out_network_specialist = result.get("office_visits", {}).get("specialist_out_network", "Not found")
        if out_network_specialist == "Not found" and out_coinsurance:
            out_network_specialist = f"{extract_percentage(out_coinsurance)}% after deductible"
        sheet[f'{out_col_letter}18'] = out_network_specialist
        
        # Urgent Care
        urgent_care = result.get("office_visits", {}).get("urgent_care", "Not found")
        urgent_care = format_benefit(urgent_care)
        sheet[f'{in_col_letter}19'] = urgent_care
        
        # Out-of-network Urgent Care
        out_network_urgent = result.get("office_visits", {}).get("urgent_care_out_network", "Not found")
        if out_network_urgent == "Not found" and out_coinsurance:
            out_network_urgent = f"{extract_percentage(out_coinsurance)}% after deductible"
        sheet[f'{out_col_letter}19'] = out_network_urgent
        
        # ----- EMERGENCY ROOM -----
        emergency_room = result.get("emergency_room", "Not found")
        # Check for special per-occurrence deductible format
        if emergency_room and "$" in str(emergency_room) and "then" in str(emergency_room).lower():
            # Already in correct format
            if "after deductible" not in str(emergency_room).lower():
                emergency_room = f"{emergency_room} after deductible"
        else:
            emergency_room = format_benefit(emergency_room)
        
        sheet[f'{in_col_letter}20'] = emergency_room
        
        # Emergency room out-of-network is always the same as in-network
        sheet[f'{out_col_letter}20'] = emergency_room
        
        # ----- PREVENTIVE SERVICES -----
        # In-network preventive is always 0%
        preventive_in = result.get("preventive_services", {}).get("in_network", "0%")
        if preventive_in == "Not found" or not preventive_in:
            preventive_in = "0%"
        sheet[f'{in_col_letter}22'] = preventive_in
        
        preventive_out = result.get("preventive_services", {}).get("out_network", "Not Covered")
        if preventive_out == "Not found" or not preventive_out:
            preventive_out = "Not Covered"
        sheet[f'{out_col_letter}22'] = preventive_out
        
        # ----- OUTPATIENT SURGERY -----
        outpatient_surgery = result.get("outpatient_surgery", "Not found")
        
        # Check for facility differences format
        if outpatient_surgery != "Not found":
            if "freestanding" in str(outpatient_surgery).lower() or "hospital" in str(outpatient_surgery).lower():
                # Already in correct format
                pass
            else:
                outpatient_surgery = format_benefit(outpatient_surgery, is_percentage=True)
        else:
            # Default to coinsurance rate if available
            if in_coinsurance and in_coinsurance != "Not found":
                outpatient_surgery = f"{extract_percentage(in_coinsurance)}% after deductible"
            else:
                outpatient_surgery = "20% after deductible"  # Standard default
                
        sheet[f'{in_col_letter}24'] = outpatient_surgery
        
        # Out-of-network outpatient surgery
        out_network_surgery = result.get("outpatient_surgery_out_network", "Not found")
        if out_network_surgery == "Not found" and out_coinsurance:
            out_network_surgery = f"{extract_percentage(out_coinsurance)}% after deductible"
        else:
            out_network_surgery = "50% after deductible"  # Standard default
            
        sheet[f'{out_col_letter}24'] = out_network_surgery
        
        # ----- INPATIENT HOSPITALIZATION -----
        hospitalization = result.get("hospitalization", "Not found")
        
        if hospitalization != "Not found":
            hospitalization = format_benefit(hospitalization, is_percentage=True)
        else:
            # Default to coinsurance rate
            if in_coinsurance and in_coinsurance != "Not found":
                hospitalization = f"{extract_percentage(in_coinsurance)}% after deductible"
            else:
                hospitalization = "20% after deductible"  # Standard default
                
        sheet[f'{in_col_letter}25'] = hospitalization
        
        # Out-of-network hospitalization
        out_network_hospitalization = result.get("hospitalization_out_network", "Not found")
        if out_network_hospitalization == "Not found" and out_coinsurance:
            out_network_hospitalization = f"{extract_percentage(out_coinsurance)}% after deductible"
        else:
            out_network_hospitalization = "50% after deductible"  # Standard default
            
        sheet[f'{out_col_letter}25'] = out_network_hospitalization
        
        # ----- IMAGING (CT/MRI) -----
        imaging = result.get("imaging", "Not found")
        
        # Check for facility differences format
        if imaging != "Not found":
            if "freestanding" in str(imaging).lower() or "hospital" in str(imaging).lower():
                # Already in correct format
                pass
            else:
                imaging = format_benefit(imaging, is_percentage=True)
        else:
            # Default to coinsurance rate if available
            if in_coinsurance and in_coinsurance != "Not found":
                imaging = f"{extract_percentage(in_coinsurance)}% after deductible"
            else:
                imaging = "20% after deductible"  # Standard default
                
        sheet[f'{in_col_letter}26'] = imaging
        
        # Out-of-network imaging
        out_network_imaging = result.get("imaging_out_network", "Not found")
        if out_network_imaging == "Not found" and out_coinsurance:
            out_network_imaging = f"{extract_percentage(out_coinsurance)}% after deductible"
        else:
            out_network_imaging = "50% after deductible"
            
        sheet[f'{out_col_letter}26'] = out_network_imaging
        
        # ----- HOSPITAL NEWBORN DELIVERY -----
        # Should match inpatient hospitalization per requirements
        sheet[f'{in_col_letter}27'] = hospitalization
        sheet[f'{out_col_letter}27'] = out_network_hospitalization
        
        # ----- PRESCRIPTION BENEFITS -----
        rx_info = result.get("prescription", {})
        
        # Rx deductible
        rx_deductible = rx_info.get("deductible", "")
        sheet[f'{in_col_letter}29'] = rx_deductible
        
        # Process each prescription tier, ensuring HSA formatting when needed
        rx_tiers = {
            "tier_1": rx_info.get("tier_1", "$10"),
            "tier_2": rx_info.get("tier_2", "$35"),
            "tier_3": rx_info.get("tier_3", "$60"),
            "tier_4": rx_info.get("tier_4", "33% up to $250"),
            "tier_5": rx_info.get("tier_5", "50% up to $500")
        }
        
        # Apply HSA formatting if needed
        if is_hsa_plan:
            for tier in rx_tiers:
                if rx_tiers[tier] and "after deductible" not in str(rx_tiers[tier]).lower():
                    rx_tiers[tier] = f"{rx_tiers[tier]} after deductible"
        
        # Fill in prescription tiers
        sheet[f'{in_col_letter}30'] = rx_tiers["tier_1"]
        sheet[f'{in_col_letter}31'] = rx_tiers["tier_2"]
        sheet[f'{in_col_letter}32'] = rx_tiers["tier_3"]
        sheet[f'{in_col_letter}33'] = rx_tiers["tier_4"]
        sheet[f'{in_col_letter}34'] = rx_tiers["tier_5"]
        
        # Mail order prescriptions - format as "Tier1 / Tier2 / Tier3"
        mail_order = rx_info.get("mail_order", "")
        
        if not mail_order:
            # Default format for mail order based on other tiers
            t1 = str(rx_tiers["tier_1"]).split(" ")[0] if rx_tiers["tier_1"] else "$10"
            t2 = str(rx_tiers["tier_2"]).split(" ")[0] if rx_tiers["tier_2"] else "$35"
            t3 = str(rx_tiers["tier_3"]).split(" ")[0] if rx_tiers["tier_3"] else "$60"
            mail_order = f"{t1} / {t2} / {t3}"
            
            if is_hsa_plan and "after deductible" not in str(mail_order).lower():
                mail_order = f"{mail_order} after deductible"
                
        sheet[f'{in_col_letter}35'] = mail_order
        
        # ----- ADDITIONAL PLAN INFO -----
        # Network information
        network_type = result.get("network_type", "PPO")
        network_name = result.get("network_name", "")
        deductible_type = result.get("deductible_type", "Embedded")
        member_website = result.get("member_website", "")
        customer_service = result.get("customer_service", "")
        
        # Fill additional information
        sheet[f'{col_letter}37'] = "2025"  # Current year as default
        sheet[f'{col_letter}38'] = "Calendar Year"  # Most common
        sheet[f'{col_letter}39'] = f"The amount you must pay for covered services before your health insurance begins to pay. {deductible_type} deductible."
        sheet[f'{col_letter}40'] = network_type
        sheet[f'{col_letter}41'] = network_name
        sheet[f'{col_letter}42'] = member_website
        sheet[f'{col_letter}43'] = customer_service
        
        # Save the workbook
        workbook.save(output_path)
        return True
    
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        return False